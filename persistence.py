from openpyxl import Workbook, load_workbook
from datetime import datetime
from typing import List
from models import AccountCategory, Entry
ope

class ExcelPersistence:
    def __init__(self, path: str):
        self.path = path

    def save(self, categories: List[AccountCategory], owner_name: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Finanças"

        ws["A1"] = f"Relatório Financeiro - {owner_name}"
        ws["A2"] = f"Data de geração: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.append(["Categoria", "Descrição", "Valor (R$)"])

        for cat in categories:
            for e in cat.entries():
                ws.append([e.categoria, e.descricao, e.valor])

        for col in ["A", "B", "C"]:
            ws.column_dimensions[col].width = 25

        wb.save(self.path)

    def load(self) -> List[Entry]:
        wb = load_workbook(self.path)
        ws = wb.active
        rows = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 3:
                continue

            categoria, descricao, valor = row

            if categoria is None:
                continue

            rows.append(Entry(categoria, descricao or "", float(valor or 0)))

        return rows
